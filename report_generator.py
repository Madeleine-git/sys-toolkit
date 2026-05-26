#!/usr/bin/env python3
"""
report_generator.py
===================
Genera informes ejecutivos en Excel a partir del inventario CSV.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


# ─── CONSTANTES ──────────────────────────────────────────────
RUTA_CSV: str   = "data/inventory.csv"
RUTA_EXCEL: str = "data/informe_servidores.xlsx"

COLOR_CABECERA: str = "1F3864"
COLOR_FILA_PAR: str = "EBF3FB"
COLOR_TITULO:   str = "2E75B6"


# ─── FUNCIÓN 1: CARGAR DATOS ─────────────────────────────────

def cargar_datos(ruta: str = RUTA_CSV) -> pd.DataFrame:
    """Carga el CSV en un DataFrame."""
    if not os.path.exists(ruta):
        print(f"  [!] No se encontró {ruta}")
        print("  [!] Ejecuta primero generate_inventory.py")
        return pd.DataFrame()

    df = pd.read_csv(ruta, encoding="utf-8")
    print(f"  ✓ CSV cargado: {len(df)} servidores")
    return df


# ─── FUNCIÓN 2: FORMATO CABECERA ─────────────────────────────

def aplicar_formato_cabecera(
    hoja: openpyxl.worksheet.worksheet.Worksheet,
    fila: int,
    num_columnas: int
) -> None:
    """Aplica formato azul oscuro a la fila de cabecera."""
    for col in range(1, num_columnas + 1):
        celda = hoja.cell(row=fila, column=col)
        celda.fill = PatternFill(
            start_color=COLOR_CABECERA,
            end_color=COLOR_CABECERA,
            fill_type="solid"
        )
        celda.font = Font(color="FFFFFF", bold=True, size=11)
        celda.alignment = Alignment(
            horizontal="center", vertical="center"
        )


def aplicar_formato_filas(
    hoja: openpyxl.worksheet.worksheet.Worksheet,
    fila_inicio: int,
    fila_fin: int,
    num_columnas: int
) -> None:
    """Aplica color alternado a las filas de datos."""
    for fila in range(fila_inicio, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            celda = hoja.cell(row=fila, column=col)
            celda.alignment = Alignment(vertical="center")
            if fila % 2 == 0:
                celda.fill = PatternFill(
                    start_color=COLOR_FILA_PAR,
                    end_color=COLOR_FILA_PAR,
                    fill_type="solid"
                )


def ajustar_anchos(
    hoja: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    """Ajusta el ancho de columnas automáticamente."""
    for col in hoja.columns:
        max_ancho: int = 0
        letra_col: str = get_column_letter(col[0].column)
        for celda in col:
            if celda.value:
                ancho_celda: int = len(str(celda.value))
                if ancho_celda > max_ancho:
                    max_ancho = ancho_celda
        hoja.column_dimensions[letra_col].width = min(
            max_ancho + 4, 40
        )


# ─── FUNCIÓN 3: GENERAR EXCEL ─────────────────────────────────

def generar_informe_excel(
    ruta_csv: str = RUTA_CSV,
    ruta_excel: str = RUTA_EXCEL
) -> None:
    """
    Genera el informe Excel completo con múltiples hojas.

    CONCEPTO — pd.ExcelWriter:
      Permite escribir múltiples hojas en un mismo fichero.
      Cada hoja es una pestaña diferente en el libro Excel.

    CONCEPTO — .to_excel():
      Exporta un DataFrame directamente a una hoja de Excel.
    """
    df = cargar_datos(ruta_csv)
    if df.empty:
        return

    fecha_fichero: str = datetime.now().strftime("%Y%m%d")
    fecha: str = datetime.now().strftime("%d/%m/%Y %H:%M")

    nombre_fichero: str = ruta_excel.replace(
        ".xlsx", f"_{fecha_fichero}.xlsx"
    )

    print(f"\n  Generando informe Excel...")

    # Filtramos vulnerables
    es_windows = df["sistema_operativo"].str.contains(
        "Windows", case=False, na=False
    )
    poca_ram = df["ram_gb"] < 4
    vulnerables = df[es_windows | poca_ram]

    # Agrupamos por departamento
    conteo_dept = (
        df.groupby("departamento")
        .size()
        .reset_index(name="total_servidores")
        .sort_values("total_servidores", ascending=False)
    )

    # Escribimos las hojas con pandas
    with pd.ExcelWriter(nombre_fichero, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="Inventario Completo",
            index=False
        )
        print("  ✓ Hoja 1: Inventario completo")

        vulnerables.to_excel(
            writer,
            sheet_name="Vulnerables",
            index=False
        )
        print(f"  ✓ Hoja 2: Vulnerables ({len(vulnerables)} servidores)")

        conteo_dept.to_excel(
            writer,
            sheet_name="Por Departamento",
            index=False
        )
        print("  ✓ Hoja 3: Por departamento")

    # Aplicamos formato con openpyxl
    wb = openpyxl.load_workbook(nombre_fichero)

    ws1 = wb["Inventario Completo"]
    aplicar_formato_cabecera(ws1, 1, len(df.columns))
    aplicar_formato_filas(ws1, 2, min(len(df) + 1, 1001), len(df.columns))
    ajustar_anchos(ws1)

    ws2 = wb["Vulnerables"]
    aplicar_formato_cabecera(ws2, 1, len(vulnerables.columns))
    aplicar_formato_filas(ws2, 2, len(vulnerables) + 1, len(vulnerables.columns))
    ajustar_anchos(ws2)

    ws3 = wb["Por Departamento"]
    aplicar_formato_cabecera(ws3, 1, len(conteo_dept.columns))
    aplicar_formato_filas(ws3, 2, len(conteo_dept) + 1, len(conteo_dept.columns))
    ajustar_anchos(ws3)

    # Añadimos hoja de resumen ejecutivo
    ws_resumen = wb.create_sheet("Resumen Ejecutivo", 0)
    ws_resumen["B2"] = "INFORME EJECUTIVO DE SERVIDORES"
    ws_resumen["B2"].font = Font(size=16, bold=True, color=COLOR_TITULO)
    ws_resumen["B3"] = f"Generado el: {fecha}"
    ws_resumen["B3"].font = Font(size=11, color="666666", italic=True)
    ws_resumen["B5"] = "RESUMEN GENERAL"
    ws_resumen["B5"].font = Font(size=12, bold=True, color=COLOR_CABECERA)

    estadisticas: list[tuple[str, str]] = [
        ("Total de servidores",         str(len(df))),
        ("Servidores activos",           str(len(df[df["estado"] == "Activo"]))),
        ("Servidores en mantenimiento",  str(len(df[df["estado"] == "Mantenimiento"]))),
        ("Servidores apagados",          str(len(df[df["estado"] == "Apagado"]))),
        ("RAM promedio",                 f"{df['ram_gb'].mean():.1f} GB"),
        ("Servidores con Windows",       str(len(df[df["sistema_operativo"].str.contains("Windows", na=False)]))),
        ("Servidores con Linux",         str(len(df[~df["sistema_operativo"].str.contains("Windows", na=False)]))),
        ("Servidores con RAM < 4GB",     str(len(df[df["ram_gb"] < 4]))),
    ]

    for i, (etiqueta, valor) in enumerate(estadisticas, start=7):
        ws_resumen[f"B{i}"] = etiqueta
        ws_resumen[f"C{i}"] = valor
        ws_resumen[f"B{i}"].font = Font(size=11)
        ws_resumen[f"C{i}"].font = Font(size=11, bold=True)

    ajustar_anchos(ws_resumen)
    wb.save(nombre_fichero)

    print(f"\n  ✓ Informe guardado en: {nombre_fichero}")
    print(f"  ✓ Hojas: Resumen Ejecutivo, Inventario Completo, Vulnerables, Por Departamento")


# ─── FUNCIÓN 4: EJECUCIÓN AUTOMÁTICA ─────────────────────────

def ejecutar_mensualmente() -> None:
    """Ejecuta el informe automáticamente cada mes."""
    import schedule
    import time

    print("\n  Modo automático activado.")
    print("  El informe se generará el día 1 de cada mes.")
    print("  Pulsa Ctrl+C para detener.\n")

    schedule.every(30).days.do(generar_informe_excel)
    generar_informe_excel()

    try:
        while True:
            schedule.run_pending()
            time.sleep(3600)
    except KeyboardInterrupt:
        print("\n  Modo automático detenido.")


# ─── MENÚ INTERACTIVO ────────────────────────────────────────

def menu_report_generator() -> None:
    """Submenú del módulo de generación de informes."""
    while True:
        print("\n  ── Generador de informes Excel ─────────────")
        print("  a.  Generar informe Excel ahora")
        print("  b.  Activar generación automática mensual")
        print("  v.  Volver al menú principal")
        print("  ────────────────────────────────────────────")

        opcion: str = input(
            "\n  Elige una opción [a/b/v]: "
        ).strip().lower()

        if opcion == "a":
            generar_informe_excel()
        elif opcion == "b":
            ejecutar_mensualmente()
        elif opcion == "v":
            break
        else:
            print("  [!] Opción no reconocida.")


if __name__ == "__main__":
    generar_informe_excel()